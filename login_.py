from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
import pytest
import openpyxl

def getData():
        excel = openpyxl.load_workbook('datalar.xlsx')
        sheet = excel["datalar"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))

        return data

@pytest.mark.parametrize("username, password", getData)
def test_login_with_empty_username(setup_and_teardown, username, password):
    driver = setup_and_teardown
    driver.get('https://www.saucedemo.com/')
    driver.maximize_window()
    sleep(2)
    
    username_input = driver.find_element(By.ID, 'user-name')
    password_input = driver.find_element(By.ID, 'password')
    login_button = driver.find_element(By.ID, 'login-button')
    
    username_input.clear()
    password_input.clear()
    login_button.click()
    sleep(2)
    
    assert "Epic sadface: Username is required" in driver.page_source, "Warning message not displayed"